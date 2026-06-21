import logging
from datetime import timedelta

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.authentication import JWTAuthentication
from api.permissions import HasResourcePermission
from django.db.models import Q
from django.utils import timezone

from api.models import Ticket
from api.serializers import TicketSerializer, TicketDetailSerializer
from api.utils import CustomPageNumberPagination, get_paginated_response


logger = logging.getLogger(__name__)

class AdminTicketListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [HasResourcePermission]
    resource = 'ticket'
    
    def get(self, request):
        filters = request.query_params.get('filters', '')
        date_range = request.query_params.get('date_range', 'all').lower()
        sort_by = request.query_params.get('sort_by', 'created_at')
        sort_desc = request.query_params.get('sort_desc', 'true').lower() == 'true'

        tickets = Ticket.objects.select_related('category').prefetch_related('attachments').all()

        if date_range == 'week':
            tickets = tickets.filter(created_at__gte=timezone.now() - timedelta(days=7))
        elif date_range == 'month':
            tickets = tickets.filter(created_at__gte=timezone.now() - timedelta(days=30))

        if filters:
            tickets = tickets.filter(
                Q(title__icontains=filters) |
                Q(public_ticket_id__icontains=filters) |
                Q(description__icontains=filters)
            )

        allowed_sort_fields = ['created_at', 'updated_at', 'status']
        if sort_by not in allowed_sort_fields:
            sort_by = 'created_at'

        order_field = f'-{sort_by}' if sort_desc else sort_by
        tickets = tickets.order_by(order_field)

        return get_paginated_response(tickets, request, TicketSerializer)
    
class AdminTicketDetailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [HasResourcePermission]
    resource = 'ticket'
    
    def get(self, request, ticket_id):
        try:
            ticket = Ticket.objects.prefetch_related(
                    'attachments',
                    'messages__attachment',
                    'resolution__attachments',
                ).get(id=ticket_id)
            serializer = TicketDetailSerializer(ticket)
            
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error retrieving ticket {ticket_id}: {str(e)}", exc_info=True)
            return Response(
                {'error': 'Failed to retrieve ticket'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    def patch(self, request, ticket_id):
        try:
            ticket = Ticket.objects.get(id=ticket_id)
        except Ticket.DoesNotExist:
            return Response({'error': 'Ticket not found'}, status=status.HTTP_404_NOT_FOUND)

        # Restrict updates to only the 'status' field
        allowed_data = {}
        if 'status' in request.data:
            allowed_data['status'] = request.data['status']

        serializer = TicketSerializer(ticket, data=allowed_data, partial=True)
        if serializer.is_valid():
            # Set resolved_at if status changes to RESOLVED manually
            if serializer.validated_data.get('status') == Ticket.Status.RESOLVED and not ticket.resolved_at:
                serializer.save(resolved_at=timezone.now())
            else:
                serializer.save()
                
            logger.info(f"Ticket {ticket.public_ticket_id} updated by {request.user.email}")
            return Response(TicketDetailSerializer(ticket).data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, ticket_id):
        try:
            ticket = Ticket.objects.prefetch_related(
                    'attachments',
                    'messages__attachment',
                    'resolution__attachments',
                ).get(id=ticket_id)
        except Ticket.DoesNotExist:
            return Response({'error': 'Ticket not found'}, status=status.HTTP_404_NOT_FOUND)

        # Delete physical files to avoid orphaned files in storage
        for attachment in ticket.attachments.all():
            if attachment.file:
                attachment.file.delete(save=False)
                
        for message in ticket.messages.all():
            if hasattr(message, 'attachment') and message.attachment and message.attachment.file:
                message.attachment.file.delete(save=False)
                
        if hasattr(ticket, 'resolution') and ticket.resolution:
            for attachment in ticket.resolution.attachments.all():
                if attachment.file:
                    attachment.file.delete(save=False)

        public_ticket_id = ticket.public_ticket_id
        ticket.delete()
        
        logger.info(f"Ticket {public_ticket_id} deleted by {request.user.email}")
        return Response({'message': 'Ticket deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class AdminTicketExportView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [HasResourcePermission]
    resource = 'ticket'
    action_override = 'export'

    def get(self, request):
        filters = request.query_params.get('filters', '')
        date_range = request.query_params.get('date_range', 'all').lower()
        sort_by = request.query_params.get('sort_by', 'created_at')
        sort_desc = request.query_params.get('sort_desc', 'true').lower() == 'true'
        status_filter = request.query_params.get('status')
        priority_filter = request.query_params.get('priority')
        category_filter = request.query_params.get('category')

        tickets = Ticket.objects.select_related('category', 'submitted_by').all()

        if date_range == 'week':
            tickets = tickets.filter(created_at__gte=timezone.now() - timedelta(days=7))
        elif date_range == 'month':
            tickets = tickets.filter(created_at__gte=timezone.now() - timedelta(days=30))

        if status_filter:
            tickets = tickets.filter(status=status_filter)

        if priority_filter:
            tickets = tickets.filter(priority=priority_filter)

        if category_filter:
            tickets = tickets.filter(category_id=category_filter)

        if filters:
            tickets = tickets.filter(
                Q(title__icontains=filters) |
                Q(public_ticket_id__icontains=filters) |
                Q(description__icontains=filters)
            )

        allowed_sort_fields = ['created_at', 'updated_at', 'status']
        if sort_by not in allowed_sort_fields:
            sort_by = 'created_at'

        order_field = f'-{sort_by}' if sort_desc else sort_by
        tickets = tickets.order_by(order_field)

        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets"

        headers = [
            "Ticket ID",
            "Title",
            "Description",
            "Category",
            "Priority",
            "Status",
            "Submitted By",
            "Anonymous?",
            "Date Created",
            "Date Resolved"
        ]
        ws.append(headers)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for ticket in tickets:
            if ticket.is_anonymous:
                submitted_by = "Anonymous"
            else:
                submitted_by = ticket.submitted_by.email if ticket.submitted_by else "None"

            created_at_str = ticket.created_at.strftime("%Y-%m-%d %H:%M:%S") if ticket.created_at else ""
            resolved_at_str = ticket.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if ticket.resolved_at else ""

            row = [
                ticket.public_ticket_id,
                ticket.title,
                ticket.description,
                ticket.category.name if ticket.category else "",
                ticket.get_priority_display(),
                ticket.get_status_display(),
                submitted_by,
                "Yes" if ticket.is_anonymous else "No",
                created_at_str,
                resolved_at_str
            ]
            ws.append(row)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = min(len(val_str), 50)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="tickets_export.xlsx"'
        wb.save(response)
        return response

